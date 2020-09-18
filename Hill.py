# задача о гладком штампе
import numpy as np

# создание окна
from tkinter import *
from tkinter.ttk import *
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment

pi = np.pi


def fbvp(c, nst, q, xADB, zADB, alphaADB, gamma, mu, sigmaADB, fi):
    for i in range(1, nst + 1):
        x1 = xADB[i, i]
        z1 = zADB[i, i]
        alpha1 = alphaADB[i, i]
        sigma1 = sigmaADB[i, i]
        for j in range(i - 1, -1, -1):
            x2 = xADB[i - 1, j]
            z2 = zADB[i - 1, j]
            alpha2 = alphaADB[i - 1, j]
            sigma2 = sigmaADB[i - 1, j]
            alpham1 = alpha1
            alpham2 = alpha2
            sigmam1 = sigma1
            sigmam2 = sigma2
            for niter in range(0, 5):
                t1 = np.tan(alpham1 + mu)
                t2 = np.tan(alpham2 - mu)
                z = (x2 - x1 + z1 * t1 - z2 * t2) / (t1 - t2)
                x = (z - z1) * t1 + x1
                a11 = 2 * sigmam1 * np.tan(fi)
                a22 = 2 * sigmam2 * np.tan(fi)
                p1 = gamma * (z - z1 - (x - x1) * np.tan(fi))
                p2 = gamma * (z - z2 + (x - x2) * np.tan(fi))
                alpha = (p1 - p2 + sigma1 - sigma2 + a11 * alpha1 + a22 * alpha2) / (a11 + a22)
                sigma = (gamma * z + q + c * 1 / np.tan(fi)) / (1 - np.sin(fi))
                alpham1 = (alpha1 + alpha) / 2
                alpham2 = (alpha2 + alpha) / 2
                sigmam1 = (sigma1 + sigma) / 2
                sigmam2 = (sigma2 + sigma) / 2
                xADB[i, j] = x
                zADB[i, j] = z
                alphaADB[i, j] = alpha
                sigmaADB[i, j] = sigma


def sbvp(nA, nst, xABC, zABC, alphaABC, gamma, mu, sigmaABC, fi):
    for i in range(0, nst):
        for j in range(1, nA):
            x1 = xABC[i + 1, j - 1]
            z1 = zABC[i + 1, j - 1]
            alpha1 = alphaABC[i + 1, j - 1]
            sigma1 = sigmaABC[i + 1, j - 1]
            x2 = xABC[i, j]
            z2 = zABC[i, j]
            alpha2 = alphaABC[i, j]
            sigma2 = sigmaABC[i, j]
            alpham1 = alpha1
            alpham2 = alpha2
            sigmam1 = sigma1
            sigmam2 = sigma2
            for niter in range(0, 5):
                t1 = np.tan(alpham1 + mu)
                t2 = np.tan(alpham2 - mu)
                z = (x1 - x2 - z1 * t1 + z2 * t2) / (t2 - t1)
                x = (z - z1) * t1 + x1
                a11 = 2 * sigmam1 * np.tan(fi)
                a22 = 2 * sigmam2 * np.tan(fi)
                p1 = gamma * (z - z1 - (x - x1) * np.tan(fi))
                p2 = gamma * (z - z2 + (x - x2) * np.tan(fi))
                alpha = (p1 - p2 + sigma1 - sigma2 + a11 * alpha1 + a22 * alpha2) / (a11 + a22)
                sigma = sigma1 - a11 * (alpha - alpha1) + p1
                alpham1 = (alpha1 + alpha) / 2
                alpham2 = (alpha2 + alpha) / 2
                sigmam1 = (sigma1 + sigma) / 2
                sigmam2 = (sigma2 + sigma) / 2
                xABC[i + 1, j] = x
                zABC[i + 1, j] = z
                alphaABC[i + 1, j] = alpha
                sigmaABC[i + 1, j] = sigma


def tbvp(nst, xOAC, zOAC, alphaOAC, gamma, mu, sigmaOAC, fi):
    for i in range(0, nst):
        x3 = xOAC[i + 1, i]
        z3 = zOAC[i + 1, i]
        alpha3 = alphaOAC[i + 1, i]
        sigma3 = sigmaOAC[i + 1, i]
        xOAC[i + 1, i + 1] = (0 - z3) * np.tan((0 + alpha3) / 2 + mu) + x3
        x2 = xOAC[i + 1, i + 1]
        zOAC[i + 1, i + 1] = 0
        alphaOAC[i + 1, i + 1] = 0
        sigmaOAC[i + 1, i + 1] = (gamma * (0 - z3 - (x2 - x3) * np.tan(fi)) + sigma3 * (
                1 - np.tan(fi) * (0 - alpha3))) / (1 + np.tan(fi) * (0 - alpha3))
        sigma2 = sigmaOAC[i + 1, i + 1]
        z2 = zOAC[i + 1, i + 1]
        alpha2 = alphaOAC[i + 1, i + 1]
        for j in range(i + 2, nst + 1):
            x1 = xOAC[j, i]
            z1 = zOAC[j, i]
            alpha1 = alphaOAC[j, i]
            sigma1 = sigmaOAC[j, i]
            alpham1 = alpha1
            alpham2 = alpha2
            sigmam1 = sigma1
            sigmam2 = sigma2
            for niter in range(0, 5):
                t1 = np.tan(alpham1 + mu)
                t2 = np.tan(alpham2 - mu)
                z = (x1 - x2 - z1 * t1 + z2 * t2) / (t2 - t1)
                x = (z - z1) * t1 + x1
                a11 = 2 * sigmam1 * np.tan(fi)
                a22 = 2 * sigmam2 * np.tan(fi)
                p1 = gamma * (z - z1 - (x - x1) * np.tan(fi))
                p2 = gamma * (z - z2 + (x - x2) * np.tan(fi))
                alpha = (p1 - p2 + sigma1 - sigma2 + a11 * alpha1 + a22 * alpha2) / (a11 + a22)
                sigma = sigma1 - a11 * (alpha - alpha1) + p1
                alpham1 = (alpha1 + alpha) / 2
                alpham2 = (alpha2 + alpha) / 2
                sigmam1 = (sigma1 + sigma) / 2
                sigmam2 = (sigma2 + sigma) / 2
                xOAC[j, i + 1] = x
                zOAC[j, i + 1] = z
                alphaOAC[j, i + 1] = alpha
                sigmaOAC[j, i + 1] = sigma
            x2 = xOAC[j, i + 1]
            z2 = zOAC[j, i + 1]
            alpha2 = alphaOAC[j, i + 1]
            sigma2 = sigmaOAC[j, i + 1]


def new_matrix(alphaABC, alphaADB, alphaOAC, nst, nA, sigmaABC, sigmaADB, sigmaOAC, xABC, xADB, xOAC, zABC,
               zADB, zOAC):
    xADB_n = np.zeros((nst + 1, nst + 1))
    zADB_n = np.zeros((nst + 1, nst + 1))
    alphaADB_n = np.zeros((nst + 1, nst + 1))
    sigmaADB_n = np.zeros((nst + 1, nst + 1))
    for i in range(0, nst + 1):
        for j in range(0, nst + 1):
            xADB_n[i, j] = xADB[i, j]
            zADB_n[i, j] = zADB[i, j]
            alphaADB_n[i, j] = alphaADB[i, j]
            sigmaADB_n[i, j] = sigmaADB[i, j]
    xABC_n = np.zeros((nst + 1, nA))
    zABC_n = np.zeros((nst + 1, nA))
    alphaABC_n = np.zeros((nst + 1, nA))
    sigmaABC_n = np.zeros((nst + 1, nA))
    for i in range(0, nst + 1):
        for j in range(0, nA):
            xABC_n[i, j] = xABC[i, j]
            zABC_n[i, j] = zABC[i, j]
            alphaABC_n[i, j] = alphaABC[i, j]
            sigmaABC_n[i, j] = sigmaABC[i, j]
    xOAC_n = np.zeros((nst + 1, nst + 1))
    zOAC_n = np.zeros((nst + 1, nst + 1))
    alphaOAC_n = np.zeros((nst + 1, nst + 1))
    sigmaOAC_n = np.zeros((nst + 1, nst + 1))
    for i in range(0, nst + 1):
        for j in range(0, nst + 1):
            xOAC_n[i, j] = xOAC[i, j]
            zOAC_n[i, j] = zOAC[i, j]
            alphaOAC_n[i, j] = alphaOAC[i, j]
            sigmaOAC_n[i, j] = sigmaOAC[i, j]
    xADB = xADB_n
    zADB = zADB_n
    alphaADB = alphaADB_n
    sigmaADB = sigmaADB_n
    xABC = xABC_n
    zABC = zABC_n
    alphaABC = alphaABC_n
    sigmaABC = sigmaABC_n
    xOAC = xOAC_n
    zOAC = zOAC_n
    alphaOAC = alphaOAC_n
    sigmaOAC = sigmaOAC_n
    return (alphaABC, alphaABC_n, alphaADB, alphaADB_n, alphaOAC, alphaOAC_n, sigmaABC, sigmaABC_n,
            sigmaADB, sigmaADB_n, sigmaOAC, sigmaOAC_n, xABC, xABC_n, xADB, xADB_n, xOAC, xOAC_n,
            zABC, zABC_n, zADB, zADB_n, zOAC, zOAC_n)


def tbvp_ocgk(alpha1, alphaOCGK, n1, sigma1, sigmaOCGK, x1, xOCGK, z1, zOCGK, mu, h, e, r, fi, gamma):
    for i in range(0, n1 - 1):
        alpham2 = alphaOCGK[i, i + 1]
        sigmam2 = sigmaOCGK[i, i + 1]
        for piter in range(0, 5):
            x3 = xOCGK[i, i + 1]
            z3 = zOCGK[i, i + 1]
            alpha3 = alphaOCGK[i, i + 1]
            sigma3 = sigmaOCGK[i, i + 1]
            t2 = np.tan(alpham2 - mu)
            zOCGK[i + 1, i + 1] = (-h + t2 * x3 - e * t2 - z3 * t2 ** 2 + np.sqrt(
                r ** 2 * t2 ** 2 + 2 * h * z3 * t2 ** 2 - h ** 2 * t2 ** 2 - z3 ** 2 * t2 ** 2 + 2 * e * h * t2
                + 2 * z3 * t2 * x3 - 2 * h * t2 * x3 - 2 * e * z3 * t2 +
                r ** 2 + 2 * e * x3 - x3 ** 2 - e ** 2)) / (-1 - t2 ** 2)
            z1 = zOCGK[i + 1, i + 1]
            xOCGK[i + 1, i + 1] = np.sqrt(r ** 2 - z1 ** 2 + 2 * z1 * h - h ** 2) + e
            x1 = xOCGK[i + 1, i + 1]
            a22 = 2 * sigmam2 * np.tan(fi)
            p2 = gamma * (z1 - z3 + (x1 - x3) * np.tan(fi))
            alphaOCGK[i + 1, i + 1] = np.arctan((x1 - x3) / (z1 - z3)) + mu
            alpha1 = alphaOCGK[i + 1, i + 1]
            sigmaOCGK[i + 1, i + 1] = sigma3 + a22 * (alpha1 - alpha3) + p2
            sigma1 = sigmaOCGK[i + 1, i + 1]
            alpham2 = (alpha1 + alpha3) / 2
            sigmam2 = (sigma1 + sigma3) / 2

        for j in range(i + 2, n1):
            x2 = xOCGK[i, j]
            z2 = zOCGK[i, j]
            alpha2 = alphaOCGK[i, j]
            sigma2 = sigmaOCGK[i, j]
            alpham1 = alpha1
            alpham2 = alpha2
            sigmam1 = sigma1
            sigmam2 = sigma2
            for niter in range(0, 5):
                t1 = np.tan(alpham1 + mu)
                t2 = np.tan(alpham2 - mu)
                z = (x1 - x2 - z1 * t1 + z2 * t2) / (t2 - t1)
                x = (z - z1) * t1 + x1
                a11 = 2 * sigmam1 * np.tan(fi)
                a22 = 2 * sigmam2 * np.tan(fi)
                p1 = gamma * (z - z1 - (x - x1) * np.tan(fi))
                p2 = gamma * (z - z2 + (x - x2) * np.tan(fi))
                alpha = (p1 - p2 + sigma1 - sigma2 + a11 * alpha1 + a22 * alpha2) / (a11 + a22)
                sigma = sigma1 - a11 * (alpha - alpha1) + p1
                alpham1 = (alpha1 + alpha) / 2
                alpham2 = (alpha2 + alpha) / 2
                sigmam1 = (sigma1 + sigma) / 2
                sigmam2 = (sigma2 + sigma) / 2
                xOCGK[i + 1, j] = x
                zOCGK[i + 1, j] = z
                alphaOCGK[i + 1, j] = alpha
                sigmaOCGK[i + 1, j] = sigma
            x1 = xOCGK[i + 1, j]
            z1 = zOCGK[i + 1, j]
            alpha1 = alphaOCGK[i + 1, j]
            sigma1 = sigmaOCGK[i + 1, j]
    return alpha1, sigma1, x1, z1


def tbvp_okl(alphaOKL, n1, sigmaOKL, xOKL, zOKL, mu, gamma, fi):
    for i in range(0, n1 - 1):
        x3 = xOKL[i + 1, i]
        z3 = zOKL[i + 1, i]
        alpha3 = alphaOKL[i + 1, i]
        sigma3 = sigmaOKL[i + 1, i]
        xOKL[i + 1, i + 1] = (0 - z3) * np.tan((0 + alpha3) / 2 + mu) + x3
        x2 = xOKL[i + 1, i + 1]
        zOKL[i + 1, i + 1] = 0
        alphaOKL[i + 1, i + 1] = 0
        sigmaOKL[i + 1, i + 1] = (gamma * (0 - z3 - (x2 - x3) * np.tan(fi)) + sigma3 * (
                1 - np.tan(fi) * (0 - alpha3))) / (
                                         1 + np.tan(fi) * (0 - alpha3))
        sigma2 = sigmaOKL[i + 1, i + 1]
        z2 = zOKL[i + 1, i + 1]
        alpha2 = alphaOKL[i + 1, i + 1]
        for j in range(i + 2, n1):
            x1 = xOKL[j, i]
            z1 = zOKL[j, i]
            alpha1 = alphaOKL[j, i]
            sigma1 = sigmaOKL[j, i]
            alpham1 = alpha1
            alpham2 = alpha2
            sigmam1 = sigma1
            sigmam2 = sigma2
            for niter in range(0, 5):
                t1 = np.tan(alpham1 + mu)
                t2 = np.tan(alpham2 - mu)
                z = (x1 - x2 - z1 * t1 + z2 * t2) / (t2 - t1)
                x = (z - z1) * t1 + x1
                a11 = 2 * sigmam1 * np.tan(fi)
                a22 = 2 * sigmam2 * np.tan(fi)
                p1 = gamma * (z - z1 - (x - x1) * np.tan(fi))
                p2 = gamma * (z - z2 + (x - x2) * np.tan(fi))
                alpha = (p1 - p2 + sigma1 - sigma2 + a11 * alpha1 + a22 * alpha2) / (a11 + a22)
                sigma = sigma1 - a11 * (alpha - alpha1) + p1
                alpham1 = (alpha1 + alpha) / 2
                alpham2 = (alpha2 + alpha) / 2
                sigmam1 = (sigma1 + sigma) / 2
                sigmam2 = (sigma2 + sigma) / 2
                xOKL[j, i + 1] = x
                zOKL[j, i + 1] = z
                alphaOKL[j, i + 1] = alpha
                sigmaOKL[j, i + 1] = sigma
            x2 = xOKL[j, i + 1]
            z2 = zOKL[j, i + 1]
            alpha2 = alphaOKL[j, i + 1]
            sigma2 = sigmaOKL[j, i + 1]


class Application(Frame):
    def __init__(self, master):
        super(Application, self).__init__(master)
        self.grid()
        self.create_widgets()

    def create_widgets(self):
        # Удельный вес
        Label(self,
              text="y ="
              ).grid(row=0, column=0)
        self.gamma_ent = Entry(self, width=5, justify=CENTER)
        self.gamma_ent.grid(row=0, column=1)
        self.gamma_ent.insert(0, "20")
        Label(self,
              text="кН/м"
              ).grid(row=0, column=2, sticky=W)
        # Угол внутреннего трения
        Label(self,
              text="fi ="
              ).grid(row=1, column=0)
        self.fi_ent = Entry(self, width=5, justify=CENTER)
        self.fi_ent.grid(row=1, column=1)
        self.fi_ent.insert(0, "30")
        Label(self,
              text="°"
              ).grid(row=1, column=2, sticky=W)
        # Удельное сцепление
        Label(self,
              text="c ="
              ).grid(row=2, column=0)
        self.c_ent = Entry(self, width=5, justify=CENTER)
        self.c_ent.grid(row=2, column=1)
        self.c_ent.insert(0, "1")
        Label(self,
              text="кПа"
              ).grid(row=2, column=2, sticky=W)
        # Размер штампа
        Label(self,
              text="b ="
              ).grid(row=3, column=0)
        self.b_ent = Entry(self, width=5, justify=CENTER)
        self.b_ent.grid(row=3, column=1)
        self.b_ent.insert(0, "1")
        Label(self,
              text="м."
              ).grid(row=3, column=2, sticky=W)
        # Радиус искривления земли
        Label(self,
              text="R ="
              ).grid(row=4, column=0)
        self.r_ent = Entry(self, width=5, justify=CENTER)
        self.r_ent.grid(row=4, column=1)
        self.r_ent.insert(0, "100")
        Label(self,
              text="м."
              ).grid(row=4, column=2, sticky=W)
        # Глубина заложения
        Label(self,
              text="h ="
              ).grid(row=5, column=0)
        self.h_ent = Entry(self, width=5, justify=CENTER)
        self.h_ent.grid(row=5, column=1)
        self.h_ent.insert(0, "100.2")
        Label(self,
              text="м."
              ).grid(row=5, column=2, sticky=W)
        # Пригрузка
        Label(self,
              text="q ="
              ).grid(row=0, column=4)
        self.q_ent = Entry(self, width=5, justify=CENTER)
        self.q_ent.grid(row=0, column=5)
        self.q_ent.insert(0, "5")
        Label(self,
              text="кПа"
              ).grid(row=0, column=6, sticky=W)
        # LAD
        Label(self,
              text="Lu ="
              ).grid(row=1, column=4)
        self.Lu_ent = Entry(self, width=5, justify=CENTER)
        self.Lu_ent.grid(row=1, column=5)
        self.Lu_ent.insert(0, "5")
        Label(self,
              text="м."
              ).grid(row=1, column=6, sticky=W)
        # Кол-во т.А
        Label(self,
              text="nA ="
              ).grid(row=2, column=4)
        self.nA_ent = Entry(self, width=5, justify=CENTER)
        self.nA_ent.grid(row=2, column=5)
        self.nA_ent.insert(0, "25")
        Label(self,
              text="шт."
              ).grid(row=2, column=6, sticky=W)
        # Кол-во шагов на AD
        Label(self,
              text="nst ="
              ).grid(row=3, column=4)
        self.nst_ent = Entry(self, width=5, justify=CENTER)
        self.nst_ent.grid(row=3, column=5)
        self.nst_ent.insert(0, "100")
        Label(self,
              text="шт."
              ).grid(row=3, column=6, sticky=W)
        # Смещение центра
        Label(self,
              text="e ="
              ).grid(row=4, column=4)
        self.e_ent = Entry(self, width=5, justify=CENTER)
        self.e_ent.grid(row=4, column=5)
        self.e_ent.insert(0, "-5")
        Label(self,
              text="м."
              ).grid(row=4, column=6, sticky=W)
        # Кнопка решить и создать таблицу
        Button(self,
               text="Решить, записать результаты в Excel",
               command=self.solver
               ).grid(row=7, column=0, columnspan=6)
        # Кнопка вывод в Excel
        """Button(self,
               text="Table",
               command=self.create_table
               ).grid(row=7, column=4, columnspan=3)"""

    def input(self):
        lp = float(self.Lu_ent.get())
        b = float(self.b_ent.get())
        gamma = float(self.gamma_ent.get())
        fi = float(self.fi_ent.get())
        c = float(self.c_ent.get())
        q = float(self.q_ent.get())
        nst = int(self.nst_ent.get())
        nA = int(self.nA_ent.get())
        r = float(self.r_ent.get())
        h = float(self.h_ent.get())
        e = float(self.e_ent.get())
        fi = np.radians(fi)
        mu = pi / 4 - fi / 2
        return b, c, e, h, lp, nA, nst, q, r, gamma, mu, fi

    def solver(self):
        b, c, e, h, lp, nA, nst, q, r, gamma, mu, fi = self.input()
        # Определяем среднее напряжение на границе AD
        sigmaAD = (q + c * 1 / np.tan(fi)) / (1 - np.sin(fi))
        print('sigmaAD =', round(sigmaAD, 3))
        xADB = np.zeros((nst + 1, nst + 1))
        zADB = np.zeros((nst + 1, nst + 1))
        alphaADB = np.zeros((nst + 1, nst + 1))
        sigmaADB = np.zeros((nst + 1, nst + 1))

        # Задаем граничные условия на границе AD
        for i in range(0, nst + 1):
            xADB[i, i] = b / 2 + lp / nst * i
            zADB[i, i] = 0
            alphaADB[i, i] = pi / 2
            sigmaADB[i, i] = sigmaAD

        # Решаем первую краевую задачу в зоне ADB
        fbvp(c, nst, q, xADB, zADB, alphaADB, gamma, mu, sigmaADB, fi)

        xABC = np.zeros((nst + 1, nA))
        zABC = np.zeros((nst + 1, nA))
        alphaABC = np.zeros((nst + 1, nA))
        sigmaABC = np.zeros((nst + 1, nA))

        # Зададим особую точку A и начальную характеристику AB веера ABC
        x1 = xADB[0, 0]
        z1 = zADB[0, 0]
        alpha1 = alphaADB[0, 0]
        sigma1 = sigmaADB[0, 0]
        for i in range(0, nA):
            x = b / 2
            z = 0
            alpha = pi / 2 - pi / 2 * i / (nA - 1)
            sigma = (q + c * 1 / np.tan(fi)) / (1 - np.sin(fi)) * np.exp((pi - 2 * alpha) * np.tan(fi))
            xABC[0, i] = x
            zABC[0, i] = z
            alphaABC[0, i] = alpha
            sigmaABC[0, i] = sigma

        # Решаем вторую краевую задачу в зоне радиального веера ABC
        for i in range(1, nst + 1):
            xABC[i, 0] = xADB[i, 0]
            zABC[i, 0] = zADB[i, 0]
            alphaABC[i, 0] = alphaADB[i, 0]
            sigmaABC[i, 0] = sigmaADB[i, 0]
        sbvp(nA, nst, xABC, zABC, alphaABC, gamma, mu, sigmaABC, fi)

        # Решаем третью краевую задачу в зоне OAC
        xOAC = np.zeros((nst + 1, nst + 1))
        zOAC = np.zeros((nst + 1, nst + 1))
        alphaOAC = np.zeros((nst + 1, nst + 1))
        sigmaOAC = np.zeros((nst + 1, nst + 1))
        for i in range(0, nst + 1):
            xOAC[i, 0] = xABC[i, nA - 1]
            zOAC[i, 0] = zABC[i, nA - 1]
            alphaOAC[i, 0] = alphaABC[i, nA - 1]
            sigmaOAC[i, 0] = sigmaABC[i, nA - 1]
        tbvp(nst, xOAC, zOAC, alphaOAC, gamma, mu, sigmaOAC, fi)

        # Отсекаем лишнее до 0
        niter = 0
        for i in range(1, nst + 1):
            if xOAC[i, i] < 0:
                itp = -xOAC[i - 1, i - 1] / (xOAC[i, i] - xOAC[i - 1, i - 1])
                niter += 1
                if niter == 2:
                    break
                nst = i
                for i in range(0, nst):
                    xADB[nst, i] = itp * (xADB[nst, i] - xADB[nst - 1, i]) + xADB[nst - 1, i]
                    zADB[nst, i] = itp * (zADB[nst, i] - zADB[nst - 1, i]) + zADB[nst - 1, i]
                    alphaADB[nst, i] = itp * (alphaADB[nst, i] - alphaADB[nst - 1, i]) + alphaADB[nst - 1, i]
                    sigmaADB[nst, i] = itp * (sigmaADB[nst, i] - sigmaADB[nst - 1, i]) + sigmaADB[nst - 1, i]
                xADB[nst, nst] = itp * (xADB[nst, nst] - xADB[nst - 1, nst - 1]) + xADB[nst - 1, nst - 1]
                zADB[nst, nst] = itp * (zADB[nst, nst] - zADB[nst - 1, nst - 1]) + zADB[nst - 1, nst - 1]
                alphaADB[nst, nst] = itp * (alphaADB[nst, nst] - alphaADB[nst - 1, nst - 1]) + alphaADB[
                    nst - 1, nst - 1]
                sigmaADB[nst, nst] = itp * (sigmaADB[nst, nst] - sigmaADB[nst - 1, nst - 1]) + sigmaADB[
                    nst - 1, nst - 1]

                for i in range(0, nA):
                    xABC[nst, i] = itp * (xABC[nst, i] - xABC[nst - 1, i]) + xABC[nst - 1, i]
                    zABC[nst, i] = itp * (zABC[nst, i] - zABC[nst - 1, i]) + zABC[nst - 1, i]
                    alphaABC[nst, i] = itp * (alphaABC[nst, i] - alphaABC[nst - 1, i]) + alphaABC[nst - 1, i]
                    sigmaABC[nst, i] = itp * (sigmaABC[nst, i] - sigmaABC[nst - 1, i]) + sigmaABC[nst - 1, i]

                for i in range(0, nst):
                    xOAC[nst, i] = itp * (xOAC[nst, i] - xOAC[nst - 1, i]) + xOAC[nst - 1, i]
                    zOAC[nst, i] = itp * (zOAC[nst, i] - zOAC[nst - 1, i]) + zOAC[nst - 1, i]
                    alphaOAC[nst, i] = itp * (alphaOAC[nst, i] - alphaOAC[nst - 1, i]) + alphaOAC[nst - 1, i]
                    sigmaOAC[nst, i] = itp * (sigmaOAC[nst, i] - sigmaOAC[nst - 1, i]) + sigmaOAC[nst - 1, i]
                xOAC[nst, nst] = itp * (xOAC[nst, nst] - xOAC[nst - 1, nst - 1]) + xOAC[nst - 1, nst - 1]
                zOAC[nst, nst] = itp * (zOAC[nst, nst] - zOAC[nst - 1, nst - 1]) + zOAC[nst - 1, nst - 1]
                alphaOAC[nst, nst] = itp * (alphaOAC[nst, nst] - alphaOAC[nst - 1, nst - 1]) + alphaOAC[
                    nst - 1, nst - 1]
                sigmaOAC[nst, nst] = itp * (sigmaOAC[nst, nst] - sigmaOAC[nst - 1, nst - 1]) + sigmaOAC[
                    nst - 1, nst - 1]

        (alphaABC, alphaABC_n, alphaADB, alphaADB_n, alphaOAC, alphaOAC_n, sigmaABC, sigmaABC_n, sigmaADB, sigmaADB_n,
         sigmaOAC, sigmaOAC_n, xABC, xABC_n, xADB, xADB_n, xOAC, xOAC_n, zABC,
         zABC_n, zADB, zADB_n, zOAC, zOAC_n) = new_matrix(
            alphaABC, alphaADB, alphaOAC, nst, nA, sigmaABC, sigmaADB, sigmaOAC, xABC, xADB, xOAC, zABC, zADB, zOAC)

        # Отсекаем лишнее до окружности
        prov = 0
        for p in range(5):
            piter = 0
            for i in range(nst + 1):
                for j in range(nA):
                    for niter in range(5):
                        if (xABC_n[i, j] - e) ** 2 + (zABC_n[i, j] - h) ** 2 < r ** 2:
                            nst = i
                            piter += 1
                            if piter == 2:
                                break
                            last_j = j
                            prov = 1
                            for i in range(nst):
                                xADB_n[nst, i] = (xADB_n[nst, i] + xADB_n[nst - 1, i]) / 2
                                zADB_n[nst, i] = (zADB_n[nst, i] + zADB_n[nst - 1, i]) / 2
                                alphaADB_n[nst, i] = (alphaADB_n[nst, i] + alphaADB_n[nst - 1, i]) / 2
                                sigmaADB_n[nst, i] = (sigmaADB_n[nst, i] + sigmaADB_n[nst - 1, i]) / 2
                            xADB_n[nst, nst] = (xADB_n[nst, nst] + xADB_n[nst - 1, nst - 1]) / 2
                            zADB_n[nst, nst] = (zADB_n[nst, nst] + zADB_n[nst - 1, nst - 1]) / 2
                            alphaADB_n[nst, nst] = (alphaADB_n[nst, nst] + alphaADB_n[nst - 1, nst - 1]) / 2
                            sigmaADB_n[nst, nst] = (sigmaADB_n[nst, nst] + sigmaADB_n[nst - 1, nst - 1]) / 2

                            for i in range(nA):
                                xABC_n[nst, i] = (xABC_n[nst, i] + xABC_n[nst - 1, i]) / 2
                                zABC_n[nst, i] = (zABC_n[nst, i] + zABC_n[nst - 1, i]) / 2
                                alphaABC_n[nst, i] = (alphaABC_n[nst, i] + alphaABC_n[nst - 1, i]) / 2
                                sigmaABC_n[nst, i] = (sigmaABC_n[nst, i] + sigmaABC_n[nst - 1, i]) / 2

                            for i in range(nst):
                                xOAC_n[nst, i] = (xOAC_n[nst, i] + xOAC_n[nst - 1, i]) / 2
                                zOAC_n[nst, i] = (zOAC_n[nst, i] + zOAC_n[nst - 1, i]) / 2
                                alphaOAC_n[nst, i] = (alphaOAC_n[nst, i] + alphaOAC_n[nst - 1, i]) / 2
                                sigmaOAC_n[nst, i] = (sigmaOAC_n[nst, i] + sigmaOAC_n[nst - 1, i]) / 2
                            xOAC_n[nst, nst] = (xOAC_n[nst, nst] + xOAC_n[nst - 1, nst - 1]) / 2
                            zOAC_n[nst, nst] = (zOAC_n[nst, nst] + zOAC_n[nst - 1, nst - 1]) / 2
                            alphaOAC_n[nst, nst] = (alphaOAC_n[nst, nst] + alphaOAC_n[nst - 1, nst - 1]) / 2
                            sigmaOAC_n[nst, nst] = (sigmaOAC_n[nst, nst] + sigmaOAC_n[nst - 1, nst - 1]) / 2

        if prov == 1:
            print("last_j", last_j)
            (alphaABC, alphaABC_n, alphaADB, alphaADB_n, alphaOAC, alphaOAC_n, sigmaABC, sigmaABC_n,
             sigmaADB, sigmaADB_n, sigmaOAC, sigmaOAC_n, xABC, xABC_n, xADB, xADB_n, xOAC,
             xOAC_n, zABC, zABC_n, zADB, zADB_n, zOAC, zOAC_n) = new_matrix(
                alphaABC, alphaADB, alphaOAC, nst, nA, sigmaABC, sigmaADB, sigmaOAC, xABC, xADB, xOAC, zABC, zADB, zOAC)

            # Решаем третью краевую задачу в зоне OCGK
            n1 = nA - last_j + nst
            xOCGK = np.zeros((n1, n1))
            zOCGK = np.zeros((n1, n1))
            alphaOCGK = np.zeros((n1, n1))
            sigmaOCGK = np.zeros((n1, n1))

            data = range(last_j, nA)
            for num, i in enumerate(data, 1):
                xOCGK[0, num - 1] = xABC[nst, i]
                zOCGK[0, num - 1] = zABC[nst, i]
                alphaOCGK[0, num - 1] = alphaABC[nst, i]
                sigmaOCGK[0, num - 1] = sigmaABC[nst, i]
                last_num = num
            data = range(last_num, n1)
            for num, i in enumerate(data, 1):
                xOCGK[0, i] = xOAC[nst, num]
                zOCGK[0, i] = zOAC[nst, num]
                alphaOCGK[0, i] = alphaOAC[nst, num]
                sigmaOCGK[0, i] = sigmaOAC[nst, num]

            tbvp_ocgk(alpha1, alphaOCGK, n1, sigma1, sigmaOCGK, x1, xOCGK, z1, zOCGK, mu,
                                                    h, e, r, fi, gamma)

            # Решаем третью краевую задачу в зоне OKL
            xOKL = np.zeros((n1, n1))
            zOKL = np.zeros((n1, n1))
            alphaOKL = np.zeros((n1, n1))
            sigmaOKL = np.zeros((n1, n1))
            for i in range(0, n1):
                xOKL[i, 0] = xOCGK[i, n1 - 1]
                zOKL[i, 0] = zOCGK[i, n1 - 1]
                alphaOKL[i, 0] = alphaOCGK[i, n1 - 1]
                sigmaOKL[i, 0] = sigmaOCGK[i, n1 - 1]
            tbvp_okl(alphaOKL, n1, sigmaOKL, xOKL, zOKL, mu, gamma, fi)

        print("Матрицы ADB")
        print(np.round(xADB, 2))
        print(np.round(zADB, 2))
        print(np.round(alphaADB, 2))
        print(np.round(sigmaADB, 2))
        print("Матрицы ABC")
        print(np.round(xABC, 2))
        print(np.round(zABC, 2))
        print(np.round(alphaABC, 2))
        print(np.round(sigmaABC, 2))
        print("Матрицы OAC")
        print(np.round(xOAC, 2))
        print(np.round(zOAC, 2))
        print(np.round(alphaOAC, 2))
        print(np.round(sigmaOAC, 2))
        if prov == 1:
            print("Матрицы OCGK")
            print(np.round(xOCGK, 2))
            print(np.round(zOCGK, 2))
            print(np.round(alphaOCGK, 2))
            print(np.round(sigmaOCGK, 2))
            print("Матрицы OKL")
            print(np.round(xOKL, 2))
            print(np.round(zOKL, 2))
            print(np.round(alphaOKL, 2))
            print(np.round(sigmaOKL, 2))

        # расчет эпюры предельного давления
        pu_x_OAC = np.zeros((nst + 1, 1))
        pu_OAC = np.zeros((nst + 1, 1))
        for i in range(nst + 1):
            pu_x_OAC[i, 0] = xOAC[i, i]
            pu_OAC[i, 0] = sigmaOAC[i, i] * (1 + np.sin(fi) * np.cos(2 * alphaOAC[i, i])) - c / np.tan(fi)
        if prov == 1:
            pu_x_OKL = np.zeros((n1, 1))
            pu_OKL = np.zeros((n1, 1))
            for i in range(n1):
                pu_x_OKL[i, 0] = xOKL[i, i]
                pu_OKL[i, 0] = sigmaOKL[i, i] * (1 + np.sin(fi) * np.cos(2 * alphaOKL[i, i])) - c / np.tan(fi)

        # создание таблицы
        fill = PatternFill(fill_type='solid',
                           start_color='c1c1c1',
                           end_color='c2c2c2')
        border = Border(left=Side(border_style='thin',
                                  color='FF000000'),
                        right=Side(border_style='thin',
                                   color='FF000000'),
                        top=Side(border_style='thin',
                                 color='FF000000'),
                        bottom=Side(border_style='thin',
                                    color='FF000000'),
                        diagonal=Side(border_style='thin',
                                      color='FF000000'),
                        diagonal_direction=0,
                        outline=Side(border_style='thin',
                                     color='FF000000'),
                        vertical=Side(border_style='thin',
                                      color='FF000000'),
                        horizontal=Side(border_style='thin',
                                        color='FF000000')
                        )
        align_center = Alignment(horizontal='center',
                                 vertical='bottom',
                                 text_rotation=0,
                                 wrap_text=False,
                                 shrink_to_fit=False,
                                 indent=0)

        # объект
        wb = Workbook()

        # активный лист
        ws = wb.active
        ws.sheet_view.zoomScale = 85

        # название страницы
        ws.title = 'Результаты'
        ws.merge_cells('A1:F1')
        ws['A1'].value = 'Матрицы ADB'
        ws['A1'].fill = fill
        ws['A1'].alignment = align_center
        column = ['i', 'j', 'x', 'z', 'alpha', 'sigma']
        for i, value in enumerate(column):
            ws.cell(row=2, column=i + 1).value = value
            ws.cell(row=2, column=i + 1).fill = fill
            ws.cell(row=2, column=i + 1).alignment = align_center
            ws.cell(row=2, column=i + 1).border = border
        k = 2
        for i in range(nst + 1):
            for j in range(i + 1):
                k += 1
                ws.cell(row=k, column=1).value = i
                ws.cell(row=k, column=1).alignment = align_center
                ws.cell(row=k, column=1).border = border
                ws.cell(row=k, column=2).value = j
                ws.cell(row=k, column=2).alignment = align_center
                ws.cell(row=k, column=2).border = border
                ws.cell(row=k, column=3).value = xADB[i, j]
                ws.cell(row=k, column=3).alignment = align_center
                ws.cell(row=k, column=3).border = border
                ws.cell(row=k, column=3).number_format = '0.00'
                ws.cell(row=k, column=4).value = zADB[i, j]
                ws.cell(row=k, column=4).alignment = align_center
                ws.cell(row=k, column=4).border = border
                ws.cell(row=k, column=4).number_format = '0.00'
                ws.cell(row=k, column=5).value = alphaADB[i, j]
                ws.cell(row=k, column=5).alignment = align_center
                ws.cell(row=k, column=5).border = border
                ws.cell(row=k, column=5).number_format = '0.00'
                ws.cell(row=k, column=6).value = sigmaADB[i, j]
                ws.cell(row=k, column=6).alignment = align_center
                ws.cell(row=k, column=6).border = border
                ws.cell(row=k, column=6).number_format = '0.00'

        ws.merge_cells('H1:M1')
        ws['H1'].value = 'Матрицы ABC'
        ws['H1'].fill = fill
        ws['H1'].alignment = align_center
        column = ['i', 'j', 'x', 'z', 'alpha', 'sigma']
        for i, value in enumerate(column):
            ws.cell(row=2, column=i + 8).value = value
            ws.cell(row=2, column=i + 8).fill = fill
            ws.cell(row=2, column=i + 8).alignment = align_center
            ws.cell(row=2, column=i + 8).border = border
        k = 2
        for i in range(nst + 1):
            for j in range(nA):
                k += 1
                ws.cell(row=k, column=8).value = i
                ws.cell(row=k, column=8).alignment = align_center
                ws.cell(row=k, column=8).border = border
                ws.cell(row=k, column=9).value = j
                ws.cell(row=k, column=9).alignment = align_center
                ws.cell(row=k, column=9).border = border
                ws.cell(row=k, column=10).value = xABC[i, j]
                ws.cell(row=k, column=10).alignment = align_center
                ws.cell(row=k, column=10).border = border
                ws.cell(row=k, column=10).number_format = '0.00'
                ws.cell(row=k, column=11).value = zABC[i, j]
                ws.cell(row=k, column=11).alignment = align_center
                ws.cell(row=k, column=11).border = border
                ws.cell(row=k, column=11).number_format = '0.00'
                ws.cell(row=k, column=12).value = alphaABC[i, j]
                ws.cell(row=k, column=12).alignment = align_center
                ws.cell(row=k, column=12).border = border
                ws.cell(row=k, column=12).number_format = '0.00'
                ws.cell(row=k, column=13).value = sigmaABC[i, j]
                ws.cell(row=k, column=13).alignment = align_center
                ws.cell(row=k, column=13).border = border
                ws.cell(row=k, column=13).number_format = '0.00'

        ws.merge_cells('O1:T1')
        ws['O1'].value = 'Матрицы OAC'
        ws['O1'].fill = fill
        ws['O1'].alignment = align_center
        column = ['i', 'j', 'x', 'z', 'alpha', 'sigma']
        for i, value in enumerate(column):
            ws.cell(row=2, column=i + 15).value = value
            ws.cell(row=2, column=i + 15).fill = fill
            ws.cell(row=2, column=i + 15).alignment = align_center
            ws.cell(row=2, column=i + 15).border = border
        k = 2
        for i in range(nst + 1):
            for j in range(i + 1):
                k += 1
                ws.cell(row=k, column=15).value = i
                ws.cell(row=k, column=15).alignment = align_center
                ws.cell(row=k, column=15).border = border
                ws.cell(row=k, column=16).value = j
                ws.cell(row=k, column=16).alignment = align_center
                ws.cell(row=k, column=16).border = border
                ws.cell(row=k, column=17).value = xOAC[i, j]
                ws.cell(row=k, column=17).alignment = align_center
                ws.cell(row=k, column=17).border = border
                ws.cell(row=k, column=17).number_format = '0.00'
                ws.cell(row=k, column=18).value = zOAC[i, j]
                ws.cell(row=k, column=18).alignment = align_center
                ws.cell(row=k, column=18).border = border
                ws.cell(row=k, column=18).number_format = '0.00'
                ws.cell(row=k, column=19).value = alphaOAC[i, j]
                ws.cell(row=k, column=19).alignment = align_center
                ws.cell(row=k, column=19).border = border
                ws.cell(row=k, column=19).number_format = '0.00'
                ws.cell(row=k, column=20).value = sigmaOAC[i, j]
                ws.cell(row=k, column=20).alignment = align_center
                ws.cell(row=k, column=20).border = border
                ws.cell(row=k, column=20).number_format = '0.00'

        if prov == 1:
            ws.merge_cells('V1:AA1')
            ws['V1'].value = 'Матрицы OCGK'
            ws['V1'].fill = fill
            ws['V1'].alignment = align_center
            column = ['i', 'j', 'x', 'z', 'alpha', 'sigma']
            for i, value in enumerate(column):
                ws.cell(row=2, column=i + 22).value = value
                ws.cell(row=2, column=i + 22).fill = fill
                ws.cell(row=2, column=i + 22).alignment = align_center
                ws.cell(row=2, column=i + 22).border = border
            k = 2
            for j in range(n1):
                for i in range(j + 1):
                    k += 1
                    ws.cell(row=k, column=22).value = i
                    ws.cell(row=k, column=22).alignment = align_center
                    ws.cell(row=k, column=22).border = border
                    ws.cell(row=k, column=23).value = j
                    ws.cell(row=k, column=23).alignment = align_center
                    ws.cell(row=k, column=23).border = border
                    ws.cell(row=k, column=24).value = xOCGK[i, j]
                    ws.cell(row=k, column=24).alignment = align_center
                    ws.cell(row=k, column=24).border = border
                    ws.cell(row=k, column=24).number_format = '0.00'
                    ws.cell(row=k, column=25).value = zOCGK[i, j]
                    ws.cell(row=k, column=25).alignment = align_center
                    ws.cell(row=k, column=25).border = border
                    ws.cell(row=k, column=25).number_format = '0.00'
                    ws.cell(row=k, column=26).value = alphaOCGK[i, j]
                    ws.cell(row=k, column=26).alignment = align_center
                    ws.cell(row=k, column=26).border = border
                    ws.cell(row=k, column=26).number_format = '0.00'
                    ws.cell(row=k, column=27).value = sigmaOCGK[i, j]
                    ws.cell(row=k, column=27).alignment = align_center
                    ws.cell(row=k, column=27).border = border
                    ws.cell(row=k, column=27).number_format = '0.00'

            ws.merge_cells('AC1:AH1')
            ws['AC1'].value = 'Матрицы OKL'
            ws['AC1'].fill = fill
            ws['AC1'].alignment = align_center
            column = ['i', 'j', 'x', 'z', 'alpha', 'sigma']
            for i, value in enumerate(column):
                ws.cell(row=2, column=i + 29).value = value
                ws.cell(row=2, column=i + 29).fill = fill
                ws.cell(row=2, column=i + 29).alignment = align_center
                ws.cell(row=2, column=i + 29).border = border
            k = 2
            for i in range(n1):
                for j in range(i + 1):
                    k += 1
                    ws.cell(row=k, column=29).value = i
                    ws.cell(row=k, column=29).alignment = align_center
                    ws.cell(row=k, column=29).border = border
                    ws.cell(row=k, column=30).value = j
                    ws.cell(row=k, column=30).alignment = align_center
                    ws.cell(row=k, column=30).border = border
                    ws.cell(row=k, column=31).value = xOKL[i, j]
                    ws.cell(row=k, column=31).alignment = align_center
                    ws.cell(row=k, column=31).border = border
                    ws.cell(row=k, column=31).number_format = '0.00'
                    ws.cell(row=k, column=32).value = zOKL[i, j]
                    ws.cell(row=k, column=32).alignment = align_center
                    ws.cell(row=k, column=32).border = border
                    ws.cell(row=k, column=32).number_format = '0.00'
                    ws.cell(row=k, column=33).value = alphaOKL[i, j]
                    ws.cell(row=k, column=33).alignment = align_center
                    ws.cell(row=k, column=33).border = border
                    ws.cell(row=k, column=33).number_format = '0.00'
                    ws.cell(row=k, column=34).value = sigmaOKL[i, j]
                    ws.cell(row=k, column=34).alignment = align_center
                    ws.cell(row=k, column=34).border = border
                    ws.cell(row=k, column=34).number_format = '0.00'
        wb.save("Results.xlsx")

        # Отрисовка графиков
        xADB_n = xADB.astype('float')
        xADB_n[xADB_n == 0] = np.nan
        xOAC_n = xOAC.astype('float')
        xOAC_n[xOAC_n == 0] = np.nan
        if prov == 1:
            xOCGK_n = xOCGK.astype('float')
            xOCGK_n[xOCGK_n == 0] = np.nan
            xOKL_n = xOKL.astype('float')
            xOKL_n[xOKL_n == 0] = np.nan
        ztADB = np.transpose(zADB)
        xtADB = np.transpose(xADB_n)
        ztABC = np.transpose(zABC)
        xtABC = np.transpose(xABC)
        ztOAC = np.transpose(zOAC)
        xtOAC = np.transpose(xOAC_n)
        if prov == 1:
            ztOCGK = np.transpose(zOCGK)
            xtOCGK = np.transpose(xOCGK_n)
            ztOKL = np.transpose(zOKL)
            xtOKL = np.transpose(xOKL_n)

        fig = plt.figure()
        ax = fig.add_subplot()
        ax.set_title('Рисунок')
        ax.plot(xtADB, -ztADB, color='blue', linewidth=0.4)
        ax.plot(xADB_n, -zADB, color='blue', linewidth=0.4)

        ax.plot(xtABC, -ztABC, color='green', linewidth=0.4)
        ax.plot(xABC, -zABC, color='green', linewidth=0.4)

        ax.plot(xtOAC, -ztOAC, color='brown', linewidth=0.4)
        ax.plot(xOAC_n, -zOAC, color='brown', linewidth=0.4)
        ax.plot(pu_x_OAC, pu_OAC / (pu_OAC[nst, 0] * 1.75), color='red', linewidth=1)

        if prov == 1:
            ax.plot(xtOCGK, -ztOCGK, color='blue', linewidth=0.4)
            ax.plot(xOCGK_n, -zOCGK, color='blue', linewidth=0.4)

            ax.plot(xtOKL, -ztOKL, color='green', linewidth=0.4)
            ax.plot(xOKL_n, -zOKL, color='green', linewidth=0.4)
            ax.plot(pu_x_OKL, pu_OKL / (pu_OAC[nst, 0] * 1.75), color='red', linewidth=1)

        ax.add_artist(plt.Circle((e, -h), r, edgecolor='black',
                                 facecolor='gray', alpha=0.6, fill=True, linewidth=2))
        ax.add_artist(plt.Rectangle((-b / 2, 0.1), b, -0.1, edgecolor='black',
                                    facecolor='gray', alpha=1, fill=True, linewidth=1))

        ax.text(b / 2, 0, 'A', fontsize=10, fontstyle='italic')
        ax.text(xADB[nst, nst], zADB[nst, nst], 'D', fontsize=10, fontstyle='italic')
        ax.text(xADB[nst, 0], -zADB[nst, 0] - 0.05, 'B', fontsize=10, fontstyle='italic')
        ax.text(xABC[nst, nA - 1] + 0.03, -zABC[nst, nA - 1], 'C', fontsize=10, fontstyle='italic')
        ax.text(xOAC[nst, nst], -zOAC[nst, nst], 'O', fontsize=10, fontstyle='italic')
        for i in range(0, nst + 1, 4):
            ax.text(pu_x_OAC[i, 0], pu_OAC[i, 0] / (pu_OAC[nst, 0] * 1.75), np.rint(pu_OAC[i, 0]), fontsize=7,
                    fontstyle='italic')

        if prov == 1:
            ax.text(xOCGK[0, 0], -zOCGK[0, 0] - 0.07, 'G', fontsize=10, fontstyle='italic')
            ax.text(xOCGK[n1 - 1, n1 - 1], -zOCGK[n1 - 1, n1 - 1] - 0.07, 'K', fontsize=10, fontstyle='italic')
            ax.text(xOKL[n1 - 1, n1 - 1], -zOKL[n1 - 1, n1 - 1], 'L', fontsize=10, fontstyle='italic')
            ax.text(pu_x_OKL[n1 - 1, 0], pu_OKL[n1 - 1, 0] / (pu_OAC[nst, 0] * 1.75) + 0.05, 'Предельное давление, МПа',
                    fontsize=7)
            """for j in range(n1):
                for i in range(j + 1):
                    ax.text(xOCGK[i, j], -zOCGK[i, j], (i, j), fontsize=5, fontstyle='italic')
            for i in range(n1):
                for j in range(i + 1):
                    ax.text(xOKL[i, j], -zOKL[i, j], (i, j), fontsize=7, fontstyle='italic')"""
            for i in range(0, n1, 8):
                ax.text(pu_x_OKL[i, 0], pu_OKL[i, 0] / (pu_OAC[nst, 0] * 1.75), np.rint(pu_OKL[i, 0]), fontsize=7,
                        fontstyle='italic')
        else:
            ax.text(pu_x_OAC[nst, 0], pu_OAC[nst, 0] / (pu_OAC[nst, 0] * 1.75) + 0.05, 'Предельное давление, МПа',
                    fontsize=7)

        ax.xaxis.set_major_locator(ticker.MultipleLocator(1))
        ax.xaxis.set_minor_locator(ticker.MultipleLocator(0.2))
        ax.yaxis.set_major_locator(ticker.MultipleLocator(1))
        ax.yaxis.set_minor_locator(ticker.MultipleLocator(0.2))
        ax.grid(which='major',
                color='gray')
        ax.minorticks_on()
        ax.grid(which='minor',
                color='gray',
                linestyle=':')
        plt.xlim(-1, 2)
        plt.ylim(-1, 1)
        plt.gca().set_aspect('equal', adjustable='box')
        plt.show()


main_window = Tk()
main_window.title("Гладкий штамп")
main_window.geometry("260x170")
main_window.resizable(0, 0)
main_window.attributes("-toolwindow", 0)
app = Application(main_window)
main_window.iconbitmap('logo.ico')
main_window.mainloop()
